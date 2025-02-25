import json
import os
import pandas as pd
import numpy as np
import configparser
import openpyxl
import datetime
import ftplib
import math
import warnings
import paramiko


def loadData(csvtoggle, csvpath, useftp, ftpserver, ftppath):
    df = pd.DataFrame()
    root_dirnames = []
    if useftp == "ftp" or useftp == "sftp":
        if useftp == "ftp":
            ftpserver.cwd("Minecraft")
            with open("../data/usercache.json", "wb") as file:
                ftpserver.retrbinary(f"RETR usercache.json", file.write)
            names = pd.DataFrame(json.load(open("../data/usercache.json", "r")))
            ftpserver.cwd("../")
            # Get directories
            root_dirnames = ftpserver.nlst(ftppath)
            ftpserver.cwd(ftppath)
        else:
            ftpserver.chdir("Minecraft")
            ftpserver.get("usercache.json", "../data/usercache.json")
            names = pd.DataFrame(json.load(open("../data/usercache.json", "r")))
            ftpserver.chdir("..")
            # Get directories
            root_dirnames = ftpserver.listdir(ftppath)
            ftpserver.chdir(ftppath)
            
        for dirname in root_dirnames:
            if dirname[-1] == ".":
                continue
            # Go to the subfolder
            if useftp == "ftp":
                ftpserver.cwd(dirname.split("/")[-1])
                filenames = ftpserver.nlst()
            else:
                ftpserver.chdir(dirname.split("/")[-1])
                filenames = ftpserver.listdir()
            
            for filename in filenames:
                if filename == "." or filename == "..":
                    continue
                print("Now processing", filename)
                
                # Download the file to process
                local_file = f"temp_{filename}"
                with open(local_file, "wb") as file:
                    if useftp == "ftp":
                        ftpserver.retrbinary(f"RETR {filename}", file.write)
                    else:
                        ftpserver.get(filename, local_file)
                
                with open(local_file, "r") as file:
                    data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                
                os.remove(local_file)
                
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
                
            if useftp == "ftp":
                ftpserver.cwd("../")  # Move back to the parent directory
            else:
                ftpserver.chdir("..")
    else:
        names_file = open('../data/usercache.json', 'r')
        names = pd.DataFrame(json.load(names_file))
        i = -1
        path = 'cobblemonplayerdata'
        for dirpath, dirnames, filenames in os.walk(path):
            if len(dirnames) > 0:
                root_dirnames = dirnames
            for filename in filenames:
                if filename == ".gitignore":
                    continue
                print("Now processing", filename)
                file = open(path + '/' + root_dirnames[i] + '/' + filename)
                data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                # Import the JSON to a Pandas DF
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
            i += 1
    # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
    df = df.fillna(0)
    if csvtoggle == "true":
        df.to_csv(csvpath)
    return df


def most_pokemons_leaderboard(df, config, type):
    # Load the Excel file
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    if type == "standard":
        sheet_name = "leaderboard2"
    elif type == "shiny":
        sheet_name = "leaderboard3"
    elif type == "legendary":
        sheet_name = "leaderboard4"
    ws = wb[sheet_name]
    i = 0
    ExcelRows = int(config['ExcelRows'])
    ExcelCols = int(config['ExcelColumns'])
    for index, row in df[0:ExcelRows*ExcelCols].iterrows():
        ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
        ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
        ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row[0])
        i += 1
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows+3, column=2, value=now.strftime(config['LastUpdated']))
    ws.cell(row=ExcelRows+4, column=2, value=config['Subtitle'])
    wb.save(file_path)


# Read config
config = configparser.ConfigParser()
config.read('cobblemon_config.ini', encoding='utf8')

# Connect to FTP if activated
ftp_server = None
if config['FTP']['UseFTP'] == "ftp":
    ftp_server = ftplib.FTP(config['FTP']['Host'], open("../username.txt", "r").read(), open("../password.txt", "r").read())
    ftp_server.encoding = "utf-8"
if config['FTP']['UseFTP'] == "sftp":
    transport = paramiko.Transport((config['FTP']['Host'], int(config['FTP']['Port'])))
    transport.connect(username=open("../username.txt", "r").read().strip(), password=open("../password.txt", "r").read().strip())
    ftp_server = paramiko.SFTPClient.from_transport(transport)

# Load the data
if config['GLOBALMATRIX']['UseCSV'] == "false":
    df = loadData(config['GLOBALMATRIX']['CreateCSV'], config['GLOBALMATRIX']['CSVPath'], config['FTP']['UseFTP'], ftp_server, config['FTP']['Path'])
else:
    df = pd.read_csv(config['GLOBALMATRIX']['CSVPath'], index_col=[0,1,2], skipinitialspace=True)

# Close the Connection
if config['FTP']['UseFTP'] == "ftp":
    ftp_server.quit()
if config['FTP']['UseFTP'] == "sftp":
    ftp_server.close()

# Prepare the counting DF
count_df = df.drop(['caughtTimestamp', 'discoveredTimestamp', 'isShiny'], level=2)
pokemons_db = pd.read_csv('Pokemon.csv')
legendary_list = pokemons_db.loc[pokemons_db['Legendary'] == True]

# Leaderboard feature
if config['LEADERBOARD']['Enable'] == "true":
    player_sum = pd.DataFrame((count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['LEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config['LEADERBOARD'], "standard")

# Shiny leaderboard feature
if config['SHINYLEADERBOARD']['Enable'] == "true":
    player_sum = pd.DataFrame(((df == "True") | (df == True)).sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['SHINYLEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config['SHINYLEADERBOARD'], "shiny")
    
# Legendary leaderboard feature
if config['LEGLEADERBOARD']['Enable'] == "true":
    legs = legendary_list['Cobblemon'].tolist()
    leg_count_df = count_df.loc[count_df.index.get_level_values(0).isin(legs)]
    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=FutureWarning)
        leg_count_df = leg_count_df.groupby(level=0).agg(lambda x: "CAUGHT" if "CAUGHT" in x.values else 0)
    #leg_count_df.to_csv("temp.csv")
    player_sum = pd.DataFrame((leg_count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['LEGLEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config['LEGLEADERBOARD'], "legendary")

print("Done!")