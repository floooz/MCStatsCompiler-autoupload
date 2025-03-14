import json
import os
import pandas as pd # type: ignore
import numpy as np # type: ignore
import configparser
import openpyxl # type: ignore
import datetime
import ftplib
import math
import warnings
import paramiko # type: ignore
import excel2img # type: ignore
import requests # type: ignore
import base64
import stat

def log_error(config, error_message):
    """Log error to error.log file with config info and credentials check"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open("error.log", "w", encoding="utf-8") as f:
        f.write(f"=== Error logged at {timestamp} ===\n\n")
        f.write(f"Error message: {error_message}\n\n")
        
        f.write("=== INPUT Configuration ===\n")
        for key in config['INPUT']:
            f.write(f"{key}: {config['INPUT'][key]}\n")

        f.write("\n=== GIT Configuration ===\n")
        for key in config['GIT']:
            print(f"{key}: {config['GIT'][key]}")
            if key != 'token':
                f.write(f"{key}: {config['GIT'][key]}\n")
            else:
                f.write(f"{key}: {config['GIT'][key][:10]}{'*'*(len(config['GIT'][key])-10)}\n")
        
        f.write("\n=== Credentials Check ===\n")
        if config['INPUT']['Mode'] in ['ftp', 'sftp']:
            username_exists = os.path.exists("username.txt")
            password_exists = os.path.exists("password.txt")
            f.write(f"username.txt exists: {username_exists}\n")
            f.write(f"password.txt exists: {password_exists}\n")
            
            if not username_exists or not password_exists:
                f.write("\nWARNING: Required credential files are missing!\n")
        else:
            f.write("No credentials required for this input mode\n")

        f.write("\n=== Librairies check ===\n")
        try:
            f.write("pandas: OK\n")
        except ImportError:
            f.write("pandas: ERROR\n")
        
        try:
            f.write("numpy: OK\n")
        except ImportError:
            f.write("numpy: ERROR\n")
        
        try:
            f.write("openpyxl: OK\n")
        except ImportError:
            f.write("openpyxl: ERROR\n")
        
        try:
            f.write("paramiko: OK\n")
        except ImportError:
            f.write("paramiko: ERROR\n")

        try:
            f.write("excel2img: OK\n")
        except ImportError:
            f.write("excel2img: ERROR\n")

        try:
            f.write("requests: OK\n")
        except ImportError:
            f.write("requests: ERROR\n")

        try:
            f.write("base64: OK\n")
        except ImportError:
            f.write("base64: ERROR\n")

        try:
            f.write("configparser: OK\n")
        except ImportError:
            f.write("configparser: ERROR\n")

        


def list_sftp_directory(sftp, path="."):
    """List contents of directory and parent directory for debugging"""
    try:
        print(f"\nContents of current directory '{path}':")
        for entry in sftp.listdir_attr(path):
            print(f"{entry.filename:30} {'<DIR>' if stat.S_ISDIR(entry.st_mode) else '<FILE>'}")
        
        parent = os.path.dirname(path) if path != "/" else "/"
        print(f"\nContents of parent directory '{parent}':")
        for entry in sftp.listdir_attr(parent):
            print(f"{entry.filename:30} {'<DIR>' if stat.S_ISDIR(entry.st_mode) else '<FILE>'}")
    except Exception as e:
        print(f"Error listing directory: {e}")

def loadVanillaData(csvtoggle, csvpath, inputmode, ftpserver, ftppath):
    df = pd.DataFrame()
    try:
        if inputmode == "ftp" or inputmode == "sftp":
            if ftppath == "root":
                ftppath_complete = "world/stats"
            else:
                ftppath_complete = ftppath + "/world/stats"
            if inputmode == "ftp":
                ftpserver.cwd(ftppath)
                with open("data/usercache/usercache.json", "wb") as file:
                    ftpserver.retrbinary(f"RETR usercache.json", file.write)
                names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
                # Go back to root
                ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
                # Get directories
                filenames = ftpserver.nlst(ftppath_complete)
                ftpserver.cwd(ftppath_complete)
            else:
                try:
                    ftpserver.chdir(ftppath)
                except IOError:
                    print(f"Failed to change to directory {ftppath}")
                    list_sftp_directory(ftpserver)
                    raise
                
                try:
                    ftpserver.get("usercache.json", "data/usercache/usercache.json")
                except IOError:
                    print("Failed to get usercache.json")
                    list_sftp_directory(ftpserver)
                    raise

                names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
                
                try:
                    current_path = ftpserver.getcwd()
                    depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
                    if depth > 0:
                        ftpserver.chdir("../" * depth)  # Return to root
                    print(f"Trying to access {ftppath_complete}")
                    filenames = ftpserver.listdir(ftppath_complete)
                    ftpserver.chdir(ftppath_complete)
                except IOError:
                    print(f"Failed to access {ftppath_complete}")
                    list_sftp_directory(ftpserver)
                    raise

            for filename in filenames:
                if filename[-1] == ".":
                    continue
                filename = filename.split("/")[-1]
                print("Now processing", filename)
                # Download the file to process
                local_file = "data/stats"+filename
                with open(local_file, "wb") as file:
                    if inputmode == "ftp":
                        ftpserver.retrbinary(f"RETR {filename}", file.write)
                    else:
                        ftpserver.get(filename, local_file)
                with open(local_file, "r") as file:
                    data = json.load(file)
                os.remove(local_file)
                
                # Import the JSON to a Pandas DF
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)
                # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
                temp_df.index = temp_df.index.str.split('.', expand=True)
                # If a stat name has a dot in it, remove the part after the dot
                if len(temp_df.index.levshape) > 3:
                    temp_df.index = temp_df.index.droplevel(3)
                    temp_df = temp_df.groupby(level=[0,1,2]).sum()
                #print(temp_df)
                #temp_df.to_csv('temp.csv')
                if df.empty:
                    df = temp_df
                else:
                    df = df.join(temp_df, how="outer")
            
            # Go back to root
            if inputmode == "ftp":
                ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            else:
                current_path = ftpserver.getcwd()
                depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
                if depth > 0:
                    ftpserver.chdir("../" * depth)
        else:
            names_file = open('data/usercache/usercache.json', 'r')
            names = pd.DataFrame(json.load(names_file))
            for filename in os.listdir('data/stats'):
                if filename == ".gitignore":
                    continue
                print("Now processing", filename)
                file = open('data/stats/' + filename)
                data = json.load(file)
                # Import the JSON to a Pandas DF
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)
                # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
                temp_df.index = temp_df.index.str.split('.', expand=True)
                # If a stat name has a dot in it, remove the part after the dot
                if len(temp_df.index.levshape) > 3:
                    temp_df.index = temp_df.index.droplevel(3)
                    temp_df = temp_df.groupby(level=[0,1,2]).sum()
                #print(temp_df)
                #temp_df.to_csv('temp.csv')
                if df.empty:
                    df = temp_df
                else:
                    df = df.join(temp_df, how="outer")
        
        # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
        df = df.fillna(0)
        if csvtoggle == "true":
            df.to_csv(csvpath)
        return df
    except Exception as e:
        log_error(config, str(e))
        raise

def loadCobblemonData(csvtoggle, csvpath, inputmode, ftpserver, ftppath):
    df = pd.DataFrame()
    try:
        if inputmode == "ftp" or inputmode == "sftp":
            root_dirnames = []
            player_count = {}  # To handle duplicate player names
                
            if ftppath == "root":
                ftppath_complete = "world/cobblemonplayerdata"
            else:
                ftppath_complete = ftppath + "/world/cobblemonplayerdata"
            if inputmode == "ftp":
                ftpserver.cwd(ftppath)
                with open("data/usercache/usercache.json", "wb") as file:
                    ftpserver.retrbinary(f"RETR usercache.json", file.write)
                names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
                # Go back to root
                ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
                # Get directories
                root_dirnames = ftpserver.nlst(ftppath_complete)
                ftpserver.cwd(ftppath_complete)
            else:
                try:
                    ftpserver.chdir(ftppath)
                except IOError:
                    print(f"Failed to change to directory {ftppath}")
                    list_sftp_directory(ftpserver)
                    raise
                
                try:
                    ftpserver.get("usercache.json", "data/usercache/usercache.json")
                except IOError:
                    print("Failed to get usercache.json")
                    list_sftp_directory(ftpserver)
                    raise

                names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
                
                try:
                    current_path = ftpserver.getcwd()
                    depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
                    if depth > 0:
                        ftpserver.chdir("../" * depth)  # Return to root
                    print(f"Trying to access {ftppath_complete}")
                    root_dirnames = ftpserver.listdir(ftppath_complete)
                    ftpserver.chdir(ftppath_complete)
                except IOError:
                    print(f"Failed to access {ftppath_complete}")
                    list_sftp_directory(ftpserver)
                    raise
                
            for dirname in root_dirnames:
                if dirname[-1] == ".":
                    continue
                # Go to the subfolder
                if inputmode == "ftp":
                    ftpserver.cwd(dirname.split("/")[-1])
                    filenames = ftpserver.nlst()
                else:
                    ftpserver.chdir(dirname.split("/")[-1])
                    filenames = ftpserver.listdir()
                
                for filename in filenames:
                    if filename == "." or filename == "..":
                        continue
                    print("Now processing", filename)
                    
                    # Download the file to process
                    local_file = "data/cobblemonplayerdata/"+filename
                    with open(local_file, "wb") as file:
                        if inputmode == "ftp":
                            ftpserver.retrbinary(f"RETR {filename}", file.write)
                        else:
                            ftpserver.get(filename, local_file)
                    
                    with open(local_file, "r") as file:
                        data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                    
                    temp_df = pd.json_normalize(data, meta_prefix=True)
                    temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                    temp_df = temp_df.transpose().iloc[:]
                    if temp_name.empty:
                        print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                        temp_name = filename[:-5]
                        player_name = temp_name
                    else:
                        player_name = temp_name.iloc[0]
                    
                    # Manage duplicates
                    if player_name in player_count:
                        player_count[player_name] += 1
                        player_name = f"{player_name}_{player_count[player_name]}"
                    else:
                        player_count[player_name] = 1
                    
                    temp_df = temp_df.rename({0: player_name}, axis=1)
                    
                    if not temp_df.empty:
                        temp_df.index = temp_df.index.str.split('.', expand=True)
                        if df.empty:
                            df = temp_df
                        else:
                            df = df.join(temp_df, how="outer")
                    else:
                        df[player_name] = np.nan
                    
                if inputmode == "ftp":
                    ftpserver.cwd("../")  # Move back to the parent directory
                else:
                    ftpserver.chdir("..")
            # Go back to root
            if inputmode == "ftp":
                ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            else:
                current_path = ftpserver.getcwd()
                depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
                if depth > 0:
                    ftpserver.chdir("../" * depth)
        else:
            names_file = open('data/usercache/usercache.json', 'r')
            names = pd.DataFrame(json.load(names_file))
            i = -1
            path = 'data/cobblemonplayerdata'
            for dirpath, dirnames, filenames in os.walk(path):
                if len(dirnames) > 0:
                    root_dirnames = dirnames
                for filename in filenames:
                    if filename == ".gitignore":
                        continue
                    print("Now processing", filename)
                    file = open(path + '/' + root_dirnames[i] + '/' + filename)
                    data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                    # Import the JSON to a Pandas DF
                    temp_df = pd.json_normalize(data, meta_prefix=True)
                    temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                    temp_df = temp_df.transpose().iloc[:]
                    
                    if temp_name.empty:
                        print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                        temp_name = filename[:-5]
                        player_name = temp_name
                    else:
                        player_name = temp_name.iloc[0]
                    
                    # Manage duplicates
                    if player_name in player_count:
                        player_count[player_name] += 1
                        player_name = f"{player_name}_{player_count[player_name]}"
                    else:
                        player_count[player_name] = 1
                    
                    temp_df = temp_df.rename({0: player_name}, axis=1)
                    
                    if not temp_df.empty:
                        temp_df.index = temp_df.index.str.split('.', expand=True)
                        if df.empty:
                            df = temp_df
                        else:
                            df = df.join(temp_df, how="outer")
                    else:
                        df[player_name] = np.nan
                i += 1
        # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
        df = df.fillna(0)
        if csvtoggle == "true":
            df.to_csv(csvpath)
        return df
    except Exception as e:
        log_error(config, str(e))
        raise

def getVanillaLeaderboard(df, cat, subcat):
    row = df.loc['stats'].loc[cat].loc[subcat].sort_values()
    print("Leaderboard of", cat, subcat, ":")
    print(row)

def getVanillaBestAndWorst(df, username, cleaning, cleaningvalue):
    if username == "null" or not username:
        print("Erreur: Aucun nom d'utilisateur spécifié dans la configuration")
        return
        
    if username not in df.columns:
        print(f"Erreur: L'utilisateur '{username}' n'existe pas dans les données")
        print("Utilisateurs disponibles:", ", ".join(df.columns))
        return
        
    nb_players = df.shape[1]
    if cleaning == "true":
        before_value = df.shape[0]
        df['zero_count'] = df.apply(lambda row: (row == 0).sum(), axis=1)
        df.drop(df[df['zero_count'] > (nb_players-int(cleaningvalue))].index, inplace=True)
        df = df.drop('zero_count', axis=1)
        print(before_value - df.shape[0], "rows dropped out of", before_value, "because of cleaning.")
    ranks = df.rank(axis=1, method='min', ascending=False)
    ranks['non_zero_values'] = df.apply(lambda row: nb_players - (row == 0).sum(), axis=1)
    ranks['value'] = df[username]
    output = ranks[[username, 'value', 'non_zero_values']].sort_values(username, ascending=False).rename(columns={username:"rank_"+username, "value":"value_"+username})
    print(output) # add .to_string() for the whole output

def most_pokemons_leaderboard(df, config, type):
    # Load the Excel file
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    if type == "standard":
        sheet_name = "leaderboard2"
    elif type == "shiny":
        sheet_name = "leaderboard3"
    elif type == "legendary":
        sheet_name = "leaderboard4"
    ws = wb[sheet_name]
    i = 0
    ExcelRows = int(config['COBBLEMONLEADERBOARDS']['ExcelRows'])
    ExcelCols = int(config['COBBLEMONLEADERBOARDS']['ExcelColumns'])
    for index, row in df[0:ExcelRows*ExcelCols].iterrows():
        ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
        ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
        ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row[0])
        i += 1
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows+3, column=2, value=now.strftime(config['COBBLEMONLEADERBOARDS']['LastUpdated']))
    ws.cell(row=ExcelRows+4, column=2, value=config['COBBLEMONLEADERBOARDS']['Subtitle'])
    wb.save(file_path)

def export_excel_to_image(config):
    """Convert Excel sheets to images"""
    
    file_path = "output.xlsx"
    # Selection of the area to export
    selection = "A1:N15"
    
    try:
        if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
        
            excel2img.export_img(
                file_path,
                "leaderboard2.png",
                "leaderboard2",
                selection
            )
    
        if config['COBBLEMONLEADERBOARDS']['ShinyEnable']== "true":
            excel2img.export_img(
                file_path,
                "leaderboard3.png",
                "leaderboard3",
                selection
            )
        
        if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
            excel2img.export_img(
                file_path,
                "leaderboard4.png",
                "leaderboard4",
                selection
            )
        
    except Exception as e:
        print("Erreur lors de l'exportation des images.")
        print(e)

def check_file_exists(api_url, headers):
    response = requests.get(api_url, headers=headers)
    return response.status_code == 200, response.json().get("sha") if response.status_code == 200 else None

def upload_image(api_url, headers, image_data):
    data = {
        "message": "Upload initial de l'image",
        "content": image_data,
        "branch": BRANCH
    }
    return requests.put(api_url, headers=headers, json=data)

def update_image(api_url, headers, image_data, sha):
    data = {
        "message": "Mise à jour de l'image",
        "content": image_data,
        "branch": BRANCH,
        "sha": sha
    }
    return requests.put(api_url, headers=headers, json=data)

# Read config
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf8')

# Connect to FTP if activated
ftp_server = None
if config['INPUT']['Mode'] == "ftp":
    try:
        ftp_server = ftplib.FTP(config['INPUT']['Host'], open("username.txt", "r").read(), open("password.txt", "r").read())
        ftp_server.encoding = "utf-8"
    except Exception as e:
        log_error(config, str(e))
        raise
if config['INPUT']['Mode'] == "sftp":
    try:
        transport = paramiko.Transport((config['INPUT']['Host'], int(config['INPUT']['Port'])))
        transport.connect(username=open("username.txt", "r").read().strip(), password=open("password.txt", "r").read().strip())
        ftp_server = paramiko.SFTPClient.from_transport(transport)
    except Exception as e:
        log_error(config, str(e))
        raise

if config['VANILLALEADERBOARD']['Enable'] == "true":
    # Load the data
    print("LOADING VANILLA DATA")
    vanilla_df = loadVanillaData(config['VANILLALEADERBOARD']['CreateCSV'], config['VANILLALEADERBOARD']['CSVPath'], config['INPUT']['Mode'], ftp_server, config['INPUT']['FTPPath'])


if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true" or config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true" or config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
    print("LOADING COBBLEMON DATA")
    if config['GLOBALMATRIX']['UseCSV'] == "false":
        cobblemon_df = loadCobblemonData(config['GLOBALMATRIX']['CreateCSV'], config['GLOBALMATRIX']['CSVPath'], config['INPUT']['Mode'], ftp_server, config['INPUT']['FTPPath'])
    else:
        cobblemon_df = pd.read_csv(config['GLOBALMATRIX']['CSVPath'], index_col=[0,1,2], skipinitialspace=True)

# Close the Connection
if config['INPUT']['Mode'] == "ftp":
    ftp_server.quit()
if config['INPUT']['Mode'] == "sftp":
    ftp_server.close()

# First leaderboard testing
if config['VANILLALEADERBOARD']['Enable'] == "true":
    getVanillaLeaderboard(vanilla_df, config['VANILLALEADERBOARD']['Category'], config['VANILLALEADERBOARD']['Subcategory'])

# First bestandworst testing
if config['BESTANDWORST']['Enable'] == "true":
    getVanillaBestAndWorst(vanilla_df, config['BESTANDWORST']['Username'], config['BESTANDWORST']['Cleaning'], config['BESTANDWORST']['CleaningValue'])

# Prepare the counting DF
count_df = cobblemon_df.drop(['caughtTimestamp', 'discoveredTimestamp', 'isShiny'], level=2)
pokemons_db = pd.read_csv('Pokemon.csv')
legendary_list = pokemons_db.loc[pokemons_db['Legendary'] == True]

# Total leaderboard feature
if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
    player_sum = pd.DataFrame((count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config, "standard")

# Shiny leaderboard feature
if config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true":
    player_sum = pd.DataFrame(((cobblemon_df == "True") | (cobblemon_df == True)).sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config, "shiny")
    
# Legendary leaderboard feature
if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
    legs = legendary_list['Cobblemon'].tolist()
    leg_count_df = count_df.loc[count_df.index.get_level_values(0).isin(legs)]
    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=FutureWarning)
        leg_count_df = leg_count_df.groupby(level=0).agg(lambda x: "CAUGHT" if "CAUGHT" in x.values else 0)
    #leg_count_df.to_csv("temp.csv")
    player_sum = pd.DataFrame((leg_count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config, "legendary")

# Export the Excel to images
export_excel_to_image(config)

if config['GIT']['UseGit'] == "true":
    GITHUB_USERNAME = config['GIT']['Username']
    REPO_NAME = config['GIT']['Repo']
    BRANCH = config['GIT']['Branch']
    GITHUB_TOKEN = config['GIT']['Token'].strip()
    
    # Vérification du token
    if not GITHUB_TOKEN or GITHUB_TOKEN.startswith('"') or GITHUB_TOKEN.endswith('"'):
        print("❌ Erreur : Le token GitHub est mal formaté. Assurez-vous qu'il n'y a pas de guillemets dans le fichier de configuration.")
        exit(1)

    try:
        if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
            with open("leaderboard2.png", "rb") as img_file:
                img_data = base64.b64encode(img_file.read()).decode("utf-8")
            headers = {
                "Authorization": f"token {GITHUB_TOKEN}",
                "Accept": "application/vnd.github.v3+json"
            }
            file_name = os.path.basename("leaderboard2.png")
            api_url = f"https://api.github.com/repos/{GITHUB_USERNAME}/{REPO_NAME}/contents/{file_name}"
            
            exists, sha = check_file_exists(api_url, headers)
            if exists:
                print("📝 Mise à jour de l'image existante...")
                response = update_image(api_url, headers, img_data, sha)
            else:
                print("⬆️ Upload d'une nouvelle image...")
                response = upload_image(api_url, headers, img_data)
            if response.status_code in [200, 201]:
                print(f"✅ Opération réussie : https://raw.githubusercontent.com/{GITHUB_USERNAME}/{REPO_NAME}/refs/heads/{BRANCH}/{file_name}")
            else:
                print(f"❌ Erreur ({response.status_code}): {response.json()}")
    except Exception as e:
        print("❌ Erreur lors de l'opération.")
        print(e)
    try:
        if config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true":
            with open("leaderboard3.png", "rb") as img_file:
                img_data = base64.b64encode(img_file.read()).decode("utf-8")
            headers = {
                "Authorization": f"token {GITHUB_TOKEN}",
                "Accept": "application/vnd.github.v3+json"
            }
            file_name = os.path.basename("leaderboard3.png")
            api_url = f"https://api.github.com/repos/{GITHUB_USERNAME}/{REPO_NAME}/contents/{file_name}"
            exists, sha = check_file_exists(api_url, headers)
            if exists:
                print("📝 Mise à jour de l'image existante...")
                response = update_image(api_url, headers, img_data, sha)
            else:
                print("⬆️ Upload d'une nouvelle image...")
                response = upload_image(api_url, headers, img_data)
            if response.status_code in [200, 201]:
                print(f"✅ Opération réussie : https://raw.githubusercontent.com/{GITHUB_USERNAME}/{REPO_NAME}/refs/heads/{BRANCH}/{file_name}")
            else:
                print(f"❌ Erreur ({response.status_code}): {response.json()}")
    except Exception as e:
        print("❌ Erreur lors de l'opération.")
        print(e)
    try:
        if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
            with open("leaderboard4.png", "rb") as img_file:
                img_data = base64.b64encode(img_file.read()).decode("utf-8")
            headers = {
                "Authorization": f"token {GITHUB_TOKEN}",
                "Accept": "application/vnd.github.v3+json"
            }
            file_name = os.path.basename("leaderboard4.png")
            api_url = f"https://api.github.com/repos/{GITHUB_USERNAME}/{REPO_NAME}/contents/{file_name}"
            exists, sha = check_file_exists(api_url, headers)
            if exists:
                print("📝 Mise à jour de l'image existante...")
                response = update_image(api_url, headers, img_data, sha)
            else:
                print("⬆️ Upload d'une nouvelle image...")
                response = upload_image(api_url, headers, img_data)
            if response.status_code in [200, 201]:
                print(f"✅ Opération réussie : https://raw.githubusercontent.com/{GITHUB_USERNAME}/{REPO_NAME}/refs/heads/{BRANCH}/{file_name}")            
            else:
                print(f"❌ Erreur ({response.status_code}): {response.json()}")
    except Exception as e:
        print("❌ Erreur lors de l'opération.")
        print(e)

print("Done!")